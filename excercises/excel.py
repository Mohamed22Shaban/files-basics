from datetime import date
import openpyxl

a =openpyxl.load_workbook(r'C:\Users\احمد شعبان\Desktop\Work Folders\first.xlsx')
print(a.sheetnames)
sheet_one = a['Sheet4']
print(sheet_one)
print(sheet_one['B3'].value)
print(sheet_one.cell(row=1,column=2).value)
print(sheet_one.max_column)   
print(sheet_one.max_row)
#=  ===========================================================
# total = 0

# for i in range(1,sheet_one.max_row +1):
#     print(sheet_one.cell(row=i,column=1).value , sheet_one.cell(row=i,column=2).value)
#     total += sheet_one.cell(row=i,column=2).value
# print(f'All Total Is {total}')


#=  ===========================================================
# to create file excel =>>>
# mu = openpyxl.Workbook()
# mu.save(filename=r'C:\Users\احمد شعبان\Desktop\Work Folders\work.xlsx')

# to change sheet's name
# fula = mu.active
# fula.title = 'first'

#to create sheet
# mu.create_sheet()

# to write
 
# l = ['mu','gmal','waleed','farg']
# d = mu['Sheet']
# for i in range(1,len(l)+1):
#     d.cell(row= 1 ,column=1).value =l[i - 1]  #=> to start from index zero


# to make chart

excel_file = openpyxl.load_workbook(r'C:\Users\احمد شعبان\Desktop\Work Folders\work.xlsx')
sheet = excel_file['Sheet']
employees = openpyxl.chart.Reference(sheet ,min_col= 1 ,max_col=1 ,min_row=1 ,max_row=6)
salery = openpyxl.chart.Reference(sheet ,min_col= 2 ,max_col=2 ,min_row=1 ,max_row=6)
chart = openpyxl.chart.BarChart()  # =>can change the type of chart  >  LineChart  <<
chart.title = 'my company'
chart.add_data(salery)
chart.set_categories(employees)
sheet.add_chart(chart ,'E8')
excel_file.save(r'C:\Users\احمد شعبان\Desktop\Work Folders\work.xlsx')



# =========================================================================================

# excersice

